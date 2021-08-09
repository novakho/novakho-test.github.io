#coding=gbk

import openpyxl
print('\\')
filename = r'D:\Desktop\worksheet1.xlsx'
print(filename)

wb=openpyxl.load_workbook(filename)
ws = wb['Sheet1']
'''
i,j=2, 4

value1 = ws.cell(i,j).value
print(value1)
ws.cell(i,j).value = 10
wb.save(filename+'22.xlsx')
value1 = ws.cell(i,j).value 
print  (value1)
'''
maxcol = ws.max_column
print(maxcol)
maxrow = ws.max_row
for i in range(1,maxrow+1):
    for j in range(1,maxcol+1):
        value2=ws.cell(i,j).value
        #print(value2)
# value1=input('value1, please   ')